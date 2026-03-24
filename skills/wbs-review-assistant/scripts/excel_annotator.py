#!/usr/bin/env python3
"""
Excel Annotator
Add review comments and conditional formatting to Excel WBS files
"""

import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


class ExcelAnnotator:
    """Add review comments and formatting to Excel WBS"""

    # Severity color mapping
    SEVERITY_COLORS = {
        "critical": "FFCCCC",  # Light red
        "major": "FFE5CC",  # Light orange
        "minor": "FFFFCC",  # Light yellow
    }

    def __init__(self, wbs_path: str, output_path: Optional[str] = None):
        """
        Initialize annotator

        Args:
            wbs_path: Path to WBS Excel file
            output_path: Path for annotated output (default: wbs_annotated_TIMESTAMP.xlsx)
        """
        self.wbs_path = Path(wbs_path)
        if not self.wbs_path.exists():
            raise FileNotFoundError(f"WBS file not found: {wbs_path}")

        # Set output path
        if output_path:
            self.output_path = Path(output_path)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_path = self.wbs_path.parent / f"wbs_annotated_{timestamp}.xlsx"

        # Load workbook
        self.workbook = load_workbook(self.wbs_path)
        self.worksheet = self.workbook.active

        # Track annotations
        self.annotations_added = 0

    def add_finding(
        self,
        row: int,
        col: int,
        finding_id: str,
        severity: str,
        issue: str,
        recommendation: str,
        requirement_ref: Optional[str] = None,
    ):
        """
        Add a review finding as an Excel comment

        Args:
            row: Row number (1-indexed)
            col: Column number (1-indexed) or letter
            finding_id: Finding ID (e.g., "CRITICAL-001")
            severity: Severity level (critical, major, minor)
            issue: Issue description
            recommendation: Recommended action
            requirement_ref: Optional requirement or hearing note reference
        """
        # Convert column letter to number if needed
        if isinstance(col, str):
            from openpyxl.utils import column_index_from_string

            col = column_index_from_string(col)

        # Get cell
        cell = self.worksheet.cell(row=row, column=col)

        # Build comment text
        comment_text = f"[{finding_id}]\n"
        comment_text += f"Severity: {severity.upper()}\n\n"
        comment_text += f"Issue: {issue}\n\n"
        if requirement_ref:
            comment_text += f"Reference: {requirement_ref}\n\n"
        comment_text += f"Recommendation: {recommendation}"

        # Add comment
        cell.comment = Comment(comment_text, "WBS Reviewer")

        # Apply conditional formatting (background color)
        if severity in self.SEVERITY_COLORS:
            fill_color = self.SEVERITY_COLORS[severity]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        self.annotations_added += 1

    def add_findings_batch(self, findings: List[Dict]):
        """
        Add multiple findings at once

        Args:
            findings: List of finding dicts with keys:
                row, col, finding_id, severity, issue, recommendation, requirement_ref
        """
        for finding in findings:
            self.add_finding(
                row=finding["row"],
                col=finding["col"],
                finding_id=finding["finding_id"],
                severity=finding["severity"],
                issue=finding["issue"],
                recommendation=finding["recommendation"],
                requirement_ref=finding.get("requirement_ref"),
            )

    def create_review_summary_sheet(self, findings: List[Dict], summary_stats: Dict):
        """
        Create a new "Review Summary" sheet with issue dashboard

        Args:
            findings: List of all findings
            summary_stats: Dict with keys: critical_count, major_count, minor_count, readiness_score
        """
        # Create new sheet
        if "Review Summary" in self.workbook.sheetnames:
            del self.workbook["Review Summary"]

        summary_sheet = self.workbook.create_sheet("Review Summary", 0)  # Insert at beginning

        # Set column widths
        summary_sheet.column_dimensions["A"].width = 15
        summary_sheet.column_dimensions["B"].width = 50
        summary_sheet.column_dimensions["C"].width = 15
        summary_sheet.column_dimensions["D"].width = 60

        # Title
        row = 1
        summary_sheet[f"A{row}"] = "WBS Review Summary"
        summary_sheet[f"A{row}"].font = Font(size=16, bold=True)
        row += 2

        # Stats
        summary_sheet[f"A{row}"] = "Review Date:"
        summary_sheet[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        summary_sheet[f"A{row}"] = "Critical Issues:"
        summary_sheet[f"B{row}"] = summary_stats.get("critical_count", 0)
        summary_sheet[f"B{row}"].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        row += 1

        summary_sheet[f"A{row}"] = "Major Issues:"
        summary_sheet[f"B{row}"] = summary_stats.get("major_count", 0)
        summary_sheet[f"B{row}"].fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        row += 1

        summary_sheet[f"A{row}"] = "Minor Issues:"
        summary_sheet[f"B{row}"] = summary_stats.get("minor_count", 0)
        summary_sheet[f"B{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        row += 1

        summary_sheet[f"A{row}"] = "Readiness Score:"
        summary_sheet[f"B{row}"] = f"{summary_stats.get('readiness_score', 0)}/100"
        row += 2

        # Findings table header
        summary_sheet[f"A{row}"] = "Finding ID"
        summary_sheet[f"B{row}"] = "Issue"
        summary_sheet[f"C{row}"] = "Severity"
        summary_sheet[f"D{row}"] = "Recommendation"

        for col in ["A", "B", "C", "D"]:
            summary_sheet[f"{col}{row}"].font = Font(bold=True)

        row += 1

        # Findings list (sorted by severity)
        severity_order = {"critical": 1, "major": 2, "minor": 3}
        sorted_findings = sorted(findings, key=lambda x: severity_order.get(x["severity"], 99))

        for finding in sorted_findings:
            summary_sheet[f"A{row}"] = finding["finding_id"]
            summary_sheet[f"B{row}"] = finding["issue"]
            summary_sheet[f"C{row}"] = finding["severity"].upper()
            summary_sheet[f"D{row}"] = finding["recommendation"]

            # Apply color to severity cell
            if finding["severity"] in self.SEVERITY_COLORS:
                fill_color = self.SEVERITY_COLORS[finding["severity"]]
                summary_sheet[f"C{row}"].fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            row += 1

    def save(self):
        """Save annotated workbook"""
        self.workbook.save(self.output_path)
        return self.output_path

    def get_stats(self) -> Dict:
        """Get annotation statistics"""
        return {"annotations_added": self.annotations_added, "output_path": str(self.output_path)}


def main():
    """CLI interface for testing Excel annotator"""
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Annotate WBS Excel with review comments")
    parser.add_argument("wbs_file", help="Path to WBS Excel file")
    parser.add_argument("--findings", "-f", required=True, help="Path to findings JSON file")
    parser.add_argument("--output", "-o", help="Output Excel path")

    args = parser.parse_args()

    # Load findings
    findings_path = Path(args.findings)
    if not findings_path.exists():
        print(f"Error: Findings file not found: {args.findings}", file=sys.stderr)
        sys.exit(1)

    findings_data = json.loads(findings_path.read_text())
    findings = findings_data.get("findings", [])

    # Create annotator
    annotator = ExcelAnnotator(args.wbs_file, args.output)

    # Add findings
    print(f"Adding {len(findings)} findings to Excel...")
    annotator.add_findings_batch(findings)

    # Create summary sheet
    summary_stats = findings_data.get("summary", {})
    annotator.create_review_summary_sheet(findings, summary_stats)

    # Save
    output_path = annotator.save()
    print(f"\nAnnotated WBS saved to: {output_path}")
    print(f"Annotations added: {annotator.annotations_added}")


if __name__ == "__main__":
    main()
