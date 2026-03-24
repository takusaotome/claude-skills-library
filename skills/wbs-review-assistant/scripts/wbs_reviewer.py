#!/usr/bin/env python3
"""
WBS Reviewer
Main review engine for analyzing WBS against requirements and generating findings
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import pandas as pd
    import yaml
    from openpyxl import load_workbook
except ImportError:
    print("Error: Required packages not installed.", file=sys.stderr)
    print("Install with: pip install pandas openpyxl pyyaml", file=sys.stderr)
    sys.exit(1)

# Import local modules
from excel_annotator import ExcelAnnotator
from requirements_parser import RequirementsParser


class WBSReviewer:
    """Main WBS review engine"""

    def __init__(
        self,
        wbs_path: str,
        requirements_path: str,
        checklist_path: Optional[str] = None,
        hearing_sheet_path: Optional[str] = None,
    ):
        """
        Initialize WBS reviewer

        Args:
            wbs_path: Path to WBS Excel file
            requirements_path: Path to requirements document
            checklist_path: Path to review checklist YAML (optional)
            hearing_sheet_path: Path to hearing notes (optional)
        """
        self.wbs_path = Path(wbs_path)
        self.requirements_path = Path(requirements_path)
        self.hearing_sheet_path = Path(hearing_sheet_path) if hearing_sheet_path else None

        # Load checklist
        if checklist_path:
            self.checklist = self._load_checklist(checklist_path)
        else:
            # Use default checklist from references
            default_checklist = Path(__file__).parent.parent / "references" / "review_checklist.yaml"
            self.checklist = self._load_checklist(str(default_checklist))

        # Parse requirements
        print("Parsing requirements document...")
        req_parser = RequirementsParser(str(self.requirements_path))
        self.requirements_data = req_parser.parse()

        # Parse hearing notes if provided
        self.hearing_notes = []
        if self.hearing_sheet_path and self.hearing_sheet_path.exists():
            print("Parsing hearing notes...")
            self.hearing_notes = self._parse_hearing_notes()

        # Load WBS
        print("Loading WBS Excel file...")
        self.wbs_df = self._load_wbs()

        # Storage for findings
        self.findings: List[Dict] = []
        self.traceability_matrix: List[Dict] = []
        self.missing_tasks: List[Dict] = []

    def _load_checklist(self, checklist_path: str) -> Dict:
        """Load review checklist YAML"""
        path = Path(checklist_path)
        if not path.exists():
            raise FileNotFoundError(f"Checklist not found: {checklist_path}")

        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    def _parse_hearing_notes(self) -> List[Dict]:
        """Parse hearing notes for decisions"""
        content = self.hearing_sheet_path.read_text(encoding="utf-8")
        decisions = []

        # Look for decision markers
        decision_patterns = [
            r"(?:決定|Decision|Decided)[\s:：](.+?)(?=\n|$)",
            r"(?:合意|Agreed|Agreement)[\s:：](.+?)(?=\n|$)",
            r"(?:承認|Approved)[\s:：](.+?)(?=\n|$)",
        ]

        for pattern in decision_patterns:
            matches = re.finditer(pattern, content, re.IGNORECASE)
            for match in matches:
                decisions.append(
                    {
                        "type": "decision",
                        "content": match.group(1).strip(),
                        "context": content[max(0, match.start() - 50) : match.end() + 100],
                    }
                )

        return decisions

    def _load_wbs(self) -> pd.DataFrame:
        """Load WBS Excel file into DataFrame"""
        # Try to detect WBS structure
        wb = load_workbook(self.wbs_path, data_only=True)
        ws = wb.active

        # Convert to DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data)

        # Use first row as header
        df.columns = df.iloc[0]
        df = df[1:]

        # Detect key columns (flexible column name matching)
        self.wbs_code_col = self._find_column(df, ["WBS", "WBSコード", "Code", "ID"])
        self.task_name_col = self._find_column(df, ["Task", "タスク名", "Name", "作業名", "Activity"])
        self.effort_col = self._find_column(df, ["Effort", "工数", "Duration", "期間", "Hours"])
        self.resource_col = self._find_column(df, ["Resource", "担当", "Assigned", "担当者"])

        return df

    def _find_column(self, df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        """Find column by matching candidate names"""
        for col in df.columns:
            if col is None:
                continue
            col_str = str(col).strip()
            for candidate in candidates:
                if candidate.lower() in col_str.lower():
                    return col
        return None

    def review(self) -> Dict:
        """
        Execute full WBS review

        Returns:
            Dict with findings, traceability_matrix, missing_tasks, summary
        """
        print("\n=== Starting WBS Review ===\n")

        # Step 1: Traceability analysis
        print("Step 1: Analyzing requirement traceability...")
        self._analyze_traceability()

        # Step 2: Structural validation
        print("Step 2: Validating WBS structure...")
        self._validate_structure()

        # Step 3: Content quality checks
        print("Step 3: Checking content quality...")
        self._check_content_quality()

        # Step 4: Hearing notes cross-check
        if self.hearing_notes:
            print("Step 4: Cross-checking hearing notes...")
            self._check_hearing_notes()

        # Step 5: Calculate summary statistics
        print("Step 5: Calculating summary statistics...")
        summary = self._calculate_summary()

        print("\n=== Review Complete ===")
        print(f"Total findings: {len(self.findings)}")
        print(f"  Critical: {summary['critical_count']}")
        print(f"  Major: {summary['major_count']}")
        print(f"  Minor: {summary['minor_count']}")
        print(f"Readiness Score: {summary['readiness_score']}/100\n")

        return {
            "findings": self.findings,
            "traceability_matrix": self.traceability_matrix,
            "missing_tasks": self.missing_tasks,
            "summary": summary,
        }

    def _analyze_traceability(self):
        """Analyze requirement to WBS task traceability"""
        requirements = self.requirements_data["requirements"]

        for req in requirements:
            req_id = req["req_id"]

            # Search WBS for this requirement ID
            mapped_tasks = self._find_wbs_tasks_for_requirement(req_id)

            # Add to traceability matrix
            self.traceability_matrix.append(
                {
                    "requirement_id": req_id,
                    "requirement_name": req["description"][:80],
                    "mapped_tasks": mapped_tasks,
                    "coverage_status": "covered" if mapped_tasks else "missing",
                }
            )

            # If no mapping found, create CRITICAL finding
            if not mapped_tasks:
                self.findings.append(
                    {
                        "finding_id": f"CRITICAL-{len(self.findings) + 1:03d}",
                        "severity": "critical",
                        "category": "missing_requirement",
                        "row": None,  # Not tied to specific row
                        "col": 1,
                        "task_name": None,
                        "issue": f"Requirement {req_id} not mapped to any WBS task",
                        "requirement_ref": req_id,
                        "recommendation": f"Add task(s) to address: {req['description'][:100]}",
                    }
                )

                # Suggest missing task
                self.missing_tasks.append(
                    {"source": req_id, "suggested_task": f"Implement {req['description'][:60]}...", "priority": "high"}
                )

    def _find_wbs_tasks_for_requirement(self, req_id: str) -> List[str]:
        """Find WBS tasks that reference a requirement ID"""
        mapped_tasks = []

        if self.task_name_col is None:
            return mapped_tasks

        for idx, row in self.wbs_df.iterrows():
            task_name = str(row.get(self.task_name_col, ""))
            wbs_code = str(row.get(self.wbs_code_col, "")) if self.wbs_code_col else f"Row{idx}"

            # Check if requirement ID appears in task description
            if req_id in task_name:
                mapped_tasks.append(wbs_code)

        return mapped_tasks

    def _validate_structure(self):
        """Validate WBS hierarchical structure"""
        if self.wbs_code_col is None:
            return

        # Check WBS numbering consistency
        prev_code = None
        for idx, row in self.wbs_df.iterrows():
            wbs_code = str(row.get(self.wbs_code_col, "")).strip()

            if not wbs_code or wbs_code == "nan":
                continue

            # Parse WBS code (e.g., "1.2.3" -> [1, 2, 3])
            try:
                code_parts = [int(x) for x in wbs_code.split(".")]
            except ValueError:
                continue  # Skip non-numeric codes

            # Check for level skipping
            if prev_code:
                if len(code_parts) > len(prev_code) + 1:
                    self.findings.append(
                        {
                            "finding_id": f"MAJOR-{len(self.findings) + 1:03d}",
                            "severity": "major",
                            "category": "structure",
                            "row": idx + 2,  # +2 for header and 0-index
                            "col": self._get_col_num(self.wbs_code_col),
                            "task_name": str(row.get(self.task_name_col, "")),
                            "issue": f"WBS code {wbs_code} skips hierarchy levels",
                            "requirement_ref": None,
                            "recommendation": f"Add intermediate level(s) between {'.'.join(map(str, prev_code))} and {wbs_code}",
                        }
                    )

            prev_code = code_parts

    def _check_content_quality(self):
        """Check task-level content quality"""
        # Check for missing effort estimates
        if self.effort_col:
            for idx, row in self.wbs_df.iterrows():
                effort = row.get(self.effort_col)
                task_name = str(row.get(self.task_name_col, ""))
                wbs_code = str(row.get(self.wbs_code_col, ""))

                if pd.isna(effort) or effort == "" or effort == 0:
                    # Check if it's a leaf task (no children)
                    if self._is_leaf_task(wbs_code):
                        self.findings.append(
                            {
                                "finding_id": f"MAJOR-{len(self.findings) + 1:03d}",
                                "severity": "major",
                                "category": "missing_effort",
                                "row": idx + 2,
                                "col": self._get_col_num(self.effort_col),
                                "task_name": task_name,
                                "issue": "Leaf task missing effort estimate",
                                "requirement_ref": None,
                                "recommendation": "Add effort estimate based on similar tasks or expert judgment",
                            }
                        )

    def _is_leaf_task(self, wbs_code: str) -> bool:
        """Check if a WBS code represents a leaf task (no children)"""
        if not wbs_code or wbs_code == "nan":
            return False

        # Check if any other task starts with this code + "."
        for idx, row in self.wbs_df.iterrows():
            other_code = str(row.get(self.wbs_code_col, ""))
            if other_code.startswith(wbs_code + "."):
                return False  # Has children

        return True

    def _check_hearing_notes(self):
        """Cross-check hearing notes against WBS"""
        for decision in self.hearing_notes:
            decision_text = decision["content"]

            # Search WBS for decision keywords
            found = False
            for idx, row in self.wbs_df.iterrows():
                task_name = str(row.get(self.task_name_col, ""))
                if any(keyword in task_name for keyword in decision_text.split()[:5]):
                    found = True
                    break

            if not found:
                self.findings.append(
                    {
                        "finding_id": f"CRITICAL-{len(self.findings) + 1:03d}",
                        "severity": "critical",
                        "category": "hearing_decision_missing",
                        "row": None,
                        "col": 1,
                        "task_name": None,
                        "issue": "Hearing note decision not reflected in WBS",
                        "requirement_ref": f"Hearing note: {decision_text[:50]}...",
                        "recommendation": f"Add/update task(s) to implement: {decision_text}",
                    }
                )

    def _calculate_summary(self) -> Dict:
        """Calculate summary statistics"""
        critical_count = sum(1 for f in self.findings if f["severity"] == "critical")
        major_count = sum(1 for f in self.findings if f["severity"] == "major")
        minor_count = sum(1 for f in self.findings if f["severity"] == "minor")

        # Calculate readiness score
        base_score = 100
        readiness_score = max(0, base_score - (critical_count * 20) - (major_count * 5) - (minor_count * 1))

        # Determine status
        if readiness_score >= 90:
            status = "Ready for baseline"
        elif readiness_score >= 70:
            status = "Needs revision"
        elif readiness_score >= 50:
            status = "Significant gaps"
        else:
            status = "Incomplete"

        return {
            "total_tasks": len(self.wbs_df),
            "critical_count": critical_count,
            "major_count": major_count,
            "minor_count": minor_count,
            "readiness_score": readiness_score,
            "status": status,
            "requirements_mapped": sum(1 for t in self.traceability_matrix if t["coverage_status"] == "covered"),
            "requirements_total": len(self.traceability_matrix),
        }

    def _get_col_num(self, col_name: str) -> int:
        """Get column number from column name"""
        if col_name is None:
            return 1
        return list(self.wbs_df.columns).index(col_name) + 1

    def export_results(self, output_dir: str):
        """Export review results to multiple formats"""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 1. Export JSON
        json_path = output_path / f"wbs_gaps_{timestamp}.json"
        results = {
            "schema_version": "1.0",
            "review_timestamp": datetime.now().isoformat(),
            "wbs_file": str(self.wbs_path.name),
            "findings": self.findings,
            "traceability_matrix": self.traceability_matrix,
            "missing_tasks": self.missing_tasks,
            "summary": self._calculate_summary(),
        }
        json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Exported JSON: {json_path}")

        # 2. Export Markdown summary
        md_path = output_path / f"wbs_review_summary_{timestamp}.md"
        md_content = self._generate_markdown_report(results)
        md_path.write_text(md_content, encoding="utf-8")
        print(f"Exported Markdown: {md_path}")

        # 3. Annotate Excel
        annotator = ExcelAnnotator(str(self.wbs_path), str(output_path / f"wbs_annotated_{timestamp}.xlsx"))
        annotator.add_findings_batch([f for f in self.findings if f["row"] is not None])
        annotator.create_review_summary_sheet(self.findings, results["summary"])
        excel_path = annotator.save()
        print(f"Exported annotated Excel: {excel_path}")

        return {"json_path": str(json_path), "markdown_path": str(md_path), "excel_path": str(excel_path)}

    def _generate_markdown_report(self, results: Dict) -> str:
        """Generate Markdown summary report"""
        summary = results["summary"]
        md = "# WBS Review Summary\n\n"
        md += f"**Review Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        md += f"**WBS File:** {self.wbs_path.name}\n"
        md += f"**Requirements:** {self.requirements_path.name}\n\n"

        md += "## Executive Summary\n\n"
        md += f"- **Critical Issues:** {summary['critical_count']}\n"
        md += f"- **Major Issues:** {summary['major_count']}\n"
        md += f"- **Minor Issues:** {summary['minor_count']}\n"
        md += f"- **Overall Readiness:** {summary['status']} ({summary['readiness_score']}/100)\n\n"

        md += "## Top Priority Findings\n\n"
        top_findings = sorted(
            self.findings, key=lambda x: {"critical": 1, "major": 2, "minor": 3}.get(x["severity"], 99)
        )[:10]

        for finding in top_findings:
            md += f"### [{finding['finding_id']}] {finding['category'].replace('_', ' ').title()}\n\n"
            if finding["row"]:
                md += f"**Location:** Row {finding['row']}\n"
            md += f"**Severity:** {finding['severity'].upper()}\n"
            md += f"**Issue:** {finding['issue']}\n"
            if finding["requirement_ref"]:
                md += f"**Reference:** {finding['requirement_ref']}\n"
            md += f"**Recommendation:** {finding['recommendation']}\n\n"

        md += "## Requirements Coverage Analysis\n\n"
        md += "| Requirement ID | Requirement Name | Mapped Tasks | Status |\n"
        md += "|----------------|------------------|--------------|--------|\n"

        for trace in self.traceability_matrix[:20]:  # Show first 20
            tasks_str = ", ".join(trace["mapped_tasks"][:3]) if trace["mapped_tasks"] else "(none)"
            status_icon = "✓" if trace["coverage_status"] == "covered" else "✗"
            md += f"| {trace['requirement_id']} | {trace['requirement_name'][:40]} | {tasks_str} | {status_icon} {trace['coverage_status']} |\n"

        if len(self.traceability_matrix) > 20:
            md += f"\n*({len(self.traceability_matrix) - 20} more requirements not shown)*\n"

        md += "\n## Missing Task Candidates\n\n"
        for idx, task in enumerate(self.missing_tasks[:10], 1):
            md += f"{idx}. **{task['suggested_task']}** (from {task['source']}) - Priority: {task['priority']}\n"

        return md


def main():
    """CLI interface"""
    import argparse

    parser = argparse.ArgumentParser(description="Review WBS against requirements")
    parser.add_argument("--wbs", required=True, help="Path to WBS Excel file")
    parser.add_argument("--requirements", required=True, help="Path to requirements document")
    parser.add_argument("--hearing-sheet", help="Path to hearing notes (optional)")
    parser.add_argument("--checklist", help="Path to custom checklist YAML (optional)")
    parser.add_argument("--output-dir", default="./wbs_review_output", help="Output directory")

    args = parser.parse_args()

    # Create reviewer
    reviewer = WBSReviewer(
        wbs_path=args.wbs,
        requirements_path=args.requirements,
        checklist_path=args.checklist,
        hearing_sheet_path=args.hearing_sheet,
    )

    # Run review
    results = reviewer.review()

    # Export results
    output_paths = reviewer.export_results(args.output_dir)

    print("\n=== Review Complete ===")
    print(f"Results exported to: {args.output_dir}")


if __name__ == "__main__":
    main()
