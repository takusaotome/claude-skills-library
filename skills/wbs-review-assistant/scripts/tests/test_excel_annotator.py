"""
Tests for excel_annotator.py
"""

from pathlib import Path

import pytest
from excel_annotator import ExcelAnnotator
from openpyxl import Workbook


@pytest.fixture
def sample_wbs_excel(tmp_path):
    """Create a sample WBS Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    # Add headers
    headers = ["WBS Code", "Task Name", "Effort (hours)", "Resource"]
    ws.append(headers)

    # Add sample tasks
    tasks = [
        ["1.0", "Project Initiation", 40, "PM"],
        ["1.1", "Kickoff Meeting", 8, "PM"],
        ["1.2", "Requirements Gathering", 32, "BA"],
        ["2.0", "Design Phase", 80, "Architect"],
        ["2.1", "System Architecture", 40, "Architect"],
        ["2.2", "Database Design", 40, "DBA"],
    ]

    for task in tasks:
        ws.append(task)

    # Save
    excel_path = tmp_path / "sample_wbs.xlsx"
    wb.save(excel_path)
    return excel_path


@pytest.fixture
def sample_findings():
    """Create sample findings for testing"""
    return [
        {
            "row": 2,
            "col": 2,
            "finding_id": "CRITICAL-001",
            "severity": "critical",
            "issue": "Missing requirement mapping",
            "recommendation": "Add requirement ID to task description",
            "requirement_ref": "REQ-001",
        },
        {
            "row": 4,
            "col": 3,
            "finding_id": "MAJOR-001",
            "severity": "major",
            "issue": "Effort estimate too high",
            "recommendation": "Break down into smaller tasks",
            "requirement_ref": None,
        },
        {
            "row": 6,
            "col": 2,
            "finding_id": "MINOR-001",
            "severity": "minor",
            "issue": "Task name not descriptive",
            "recommendation": "Clarify task name with action verb",
            "requirement_ref": None,
        },
    ]


class TestExcelAnnotator:
    """Test cases for ExcelAnnotator"""

    def test_initialize_annotator(self, sample_wbs_excel):
        """Test annotator initialization"""
        annotator = ExcelAnnotator(str(sample_wbs_excel))

        assert annotator.wbs_path.exists()
        assert annotator.workbook is not None
        assert annotator.worksheet is not None
        assert annotator.annotations_added == 0

    def test_add_single_finding(self, sample_wbs_excel, tmp_path):
        """Test adding a single finding"""
        output_path = tmp_path / "annotated.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        # Add finding
        annotator.add_finding(
            row=2,
            col=2,
            finding_id="CRITICAL-001",
            severity="critical",
            issue="Test issue",
            recommendation="Test recommendation",
            requirement_ref="REQ-001",
        )

        assert annotator.annotations_added == 1

        # Check cell has comment
        cell = annotator.worksheet.cell(row=2, column=2)
        assert cell.comment is not None
        assert "CRITICAL-001" in cell.comment.text
        assert "Test issue" in cell.comment.text

        # Check cell has background color
        assert cell.fill.start_color.rgb == "FFCCCC"  # Critical = red

    def test_add_findings_batch(self, sample_wbs_excel, sample_findings, tmp_path):
        """Test adding multiple findings at once"""
        output_path = tmp_path / "annotated_batch.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        # Add findings
        annotator.add_findings_batch(sample_findings)

        assert annotator.annotations_added == 3

    def test_severity_colors(self, sample_wbs_excel, tmp_path):
        """Test that different severities get correct colors"""
        output_path = tmp_path / "annotated_colors.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        # Add findings with different severities
        annotator.add_finding(2, 1, "CRIT-1", "critical", "Issue", "Fix")
        annotator.add_finding(3, 1, "MAJ-1", "major", "Issue", "Fix")
        annotator.add_finding(4, 1, "MIN-1", "minor", "Issue", "Fix")

        # Check colors
        assert annotator.worksheet.cell(row=2, column=1).fill.start_color.rgb == "FFCCCC"  # Red
        assert annotator.worksheet.cell(row=3, column=1).fill.start_color.rgb == "FFE5CC"  # Orange
        assert annotator.worksheet.cell(row=4, column=1).fill.start_color.rgb == "FFFFCC"  # Yellow

    def test_create_review_summary_sheet(self, sample_wbs_excel, sample_findings, tmp_path):
        """Test creating review summary sheet"""
        output_path = tmp_path / "annotated_summary.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        summary_stats = {"critical_count": 1, "major_count": 1, "minor_count": 1, "readiness_score": 75}

        annotator.create_review_summary_sheet(sample_findings, summary_stats)

        # Check summary sheet exists
        assert "Review Summary" in annotator.workbook.sheetnames

        # Check summary sheet is first
        assert annotator.workbook.sheetnames[0] == "Review Summary"

        summary_sheet = annotator.workbook["Review Summary"]

        # Check title exists
        assert summary_sheet["A1"].value == "WBS Review Summary"

        # Check stats present
        assert summary_sheet["B4"].value == 1  # Critical count
        assert summary_sheet["B5"].value == 1  # Major count
        assert summary_sheet["B6"].value == 1  # Minor count
        assert summary_sheet["B7"].value == "75/100"  # Readiness score

    def test_save_annotated_file(self, sample_wbs_excel, tmp_path):
        """Test saving annotated workbook"""
        output_path = tmp_path / "saved_annotated.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        annotator.add_finding(2, 1, "TEST-1", "minor", "Test", "Fix")

        saved_path = annotator.save()

        assert saved_path.exists()
        assert saved_path == output_path

    def test_default_output_path(self, sample_wbs_excel):
        """Test that default output path includes timestamp"""
        annotator = ExcelAnnotator(str(sample_wbs_excel))

        # Check default path has timestamp pattern
        assert "wbs_annotated_" in annotator.output_path.name
        assert annotator.output_path.suffix == ".xlsx"

    def test_column_letter_support(self, sample_wbs_excel, tmp_path):
        """Test that column can be specified as letter (e.g., 'B') or number"""
        output_path = tmp_path / "col_letter.xlsx"
        annotator = ExcelAnnotator(str(sample_wbs_excel), str(output_path))

        # Add finding with column letter
        annotator.add_finding(2, "B", "TEST-1", "minor", "Test", "Fix")

        assert annotator.annotations_added == 1

        # Check comment added to column B (column 2)
        cell = annotator.worksheet.cell(row=2, column=2)
        assert cell.comment is not None

    def test_get_stats(self, sample_wbs_excel, sample_findings):
        """Test getting annotation statistics"""
        annotator = ExcelAnnotator(str(sample_wbs_excel))
        annotator.add_findings_batch(sample_findings)

        stats = annotator.get_stats()

        assert stats["annotations_added"] == 3
        assert "output_path" in stats

    def test_file_not_found(self):
        """Test handling of non-existent WBS file"""
        with pytest.raises(FileNotFoundError):
            ExcelAnnotator("/nonexistent/wbs.xlsx")
